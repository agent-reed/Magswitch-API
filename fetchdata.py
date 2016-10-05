import openpyxl  # Need this for working with excel docs
import re

def getUnitData(unit, type):

	#Open our workbook and filter to the sheet coresponding to the name of the provided unit
	wb = openpyxl.load_workbook('New Master Test Collection Document.xlsx', data_only=True) #data_only to avoid excel refrences

	try:
		sheet = wb.get_sheet_by_name(unit)
	except Exception, e:
		raise e
	data = []

	#Depending on what data we want to return, either grab the forces or thicknesses
	if type == 'thicknesses':
		for i in range(4, 24, 1):
			data.append(sheet.cell(row=i, column=2).value) #Column 2 is standard for thicknesses

	elif type == 'forces':
		for i in range(4, 24, 1):
			data.append(sheet.cell(row=i, column=9).value) # Column 9 is standard for forces

	else:
		print("Unknown type of unit data %s" %type)

	return data

def getDerekData(thickness, width, length, typeOfSteel, condition,orientation):

	wb = openpyxl.load_workbook('Lifting_Calculations.xlsx', data_only=True)
	sheet = wb.get_sheet_by_name('Sheet Sizes')

	tools = []

	for row in sheet.iter_rows():
		name = str(row[0].value)
		searchString = "(?<!.)" + re.escape(thickness + "\"") + re.escape(width + "\"") + re.escape(length + "\"") #Regex101.com is your friend
		match = re.search(searchString, name)
		if match is not None & (row[17].value == 'Yes'):
			for item in row:
				print(item.value)
			return {"Plate Weight":row[4].value,"3:1 SWL Per Magnet":row[6].value, "Saftey Factor":row[10].value, "Number of Magnets":row[8].value, "Destack":row[17].value, "Length":row[3].value, "Width":row[2].value, "Thickness":row[1].value, "Name":row[5].value }
